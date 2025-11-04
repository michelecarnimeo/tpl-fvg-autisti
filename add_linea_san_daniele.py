#!/usr/bin/env python3
# Script per aggiungere la Linea San Daniele al database.json

from openpyxl import load_workbook
import json

# Leggi il database esistente
with open('database.json', 'r', encoding='utf-8') as f:
    database = json.load(f)

print(f"Linee esistenti nel database: {len(database)}")
for i, linea in enumerate(database, 1):
    print(f"  {i}. {linea['nome']} - {len(linea['fermate'])} fermate")

# Leggi il file Excel
wb = load_workbook('Udine San Daniele.xlsx')
ws = wb.active

# Leggi la prima riga (fermate)
first_row = ws[1]
fermate = []
for cell in first_row:
    value = cell.value
    if value is not None and str(value).strip():
        fermate.append(str(value).strip())

print(f"\nFermate trovate nell'Excel: {len(fermate)}")
print(f"Prima fermata: {fermate[0]}")
print(f"Ultima fermata: {fermate[-1]}")

# Leggi la matrice dei codici (righe 2+)
codici = []
for i in range(2, ws.max_row + 1):
    row = ws[i]
    codici_riga = []
    for cell in row:
        value = cell.value
        if value is None or str(value).strip() == '':
            codici_riga.append('')
        else:
            codici_riga.append(str(value).strip())
    codici.append(codici_riga)

print(f"\nMatrice codici: {len(codici)} righe x {len(codici[0]) if codici else 0} colonne")

# Mappatura codici -> prezzi (stessa logica della Linea 400)
codici_prezzi = {
    'U1': 1.50,
    'E1': 2.50,
    'E2': 3.50,
    'E3': 4.00,
    'E4': 4.50,
    'E5': 5.50,
    'E6': 6.50,
    'E7': 7.50
}

# Genera i prezzi dalla matrice codici
num_fermate = len(fermate)
prezzi = []

for i in range(num_fermate):
    riga_prezzi = []
    for j in range(num_fermate):
        if i == j:
            riga_prezzi.append(0)
        else:
            # Prendi il codice dalla matrice (salta prima colonna se contiene nome fermata)
            if i < len(codici):
                # La prima colonna può essere il nome della fermata, quindi prendiamo da colonna j+1
                # Ma se j+1 > len(codici[i]), allora prendiamo j
                if j < len(codici[i]):
                    codice = codici[i][j] if codici[i][j] else ''
                else:
                    codice = ''
            else:
                codice = ''
            
            if codice and codice in codici_prezzi:
                riga_prezzi.append(codici_prezzi[codice])
            else:
                riga_prezzi.append(0)
    prezzi.append(riga_prezzi)

print(f"\nPrezzi generati: {len(prezzi)} righe x {len(prezzi[0]) if prezzi else 0} colonne")

# Crea l'oggetto linea
nuova_linea = {
    "nome": "Linea 401 Udine-San Daniele",
    "fermate": fermate,
    "codici": codici,
    "prezzi": prezzi
}

# Aggiungi al database
database.append(nuova_linea)

# Salva il database aggiornato
with open('database.json', 'w', encoding='utf-8') as f:
    json.dump(database, f, indent=2, ensure_ascii=False)

print(f"\n✅ Linea aggiunta con successo!")
print(f"   Nome: {nuova_linea['nome']}")
print(f"   Fermate: {len(nuova_linea['fermate'])}")
print(f"   Codici: matrice {len(codici)}x{len(codici[0]) if codici else 0}")
print(f"   Prezzi: matrice {len(prezzi)}x{len(prezzi[0]) if prezzi else 0}")
print(f"\nTotale linee nel database: {len(database)}")

