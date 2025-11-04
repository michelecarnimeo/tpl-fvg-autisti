#!/usr/bin/env python3
# Script per leggere file Excel e convertire in formato database.json

from openpyxl import load_workbook
import json

wb = load_workbook('Udine San Daniele.xlsx')
ws = wb.active

# Leggi la prima riga (nomi delle fermate)
fermate = []
first_row = ws[1]
for cell in first_row:
    if cell.value and str(cell.value).strip():
        fermate.append(str(cell.value).strip())

# Rimuovi la prima cella se è vuota o contiene "Udine" come header
if fermate and (not fermate[0] or fermate[0] == "Udine"):
    fermate = fermate[1:]

print(f"Numero fermate trovate: {len(fermate)}")
print(f"Prime 5 fermate: {fermate[:5]}")
print(f"Ultime 5 fermate: {fermate[-5:]}")

# Leggi la matrice dei codici (righe 2+)
codici = []
for i in range(2, ws.max_row + 1):
    row = ws[i]
    codici_riga = []
    # Salta la prima cella (nome fermata)
    for j, cell in enumerate(row[1:], 1):
        value = cell.value
        if value is None or str(value).strip() == '':
            codici_riga.append('')
        else:
            codici_riga.append(str(value).strip())
    codici.append(codici_riga)

print(f"\nMatrice codici: {len(codici)} righe x {len(codici[0]) if codici else 0} colonne")

# Verifica se ci sono prezzi (cerca colonne numeriche)
print("\nVerifica presenza prezzi...")
has_prices = False
# Controlla se ci sono valori numerici oltre ai codici

# Genera l'oggetto linea
linea_data = {
    "nome": "Linea Udine - San Daniele",
    "fermate": fermate,
    "codici": codici
}

print("\n" + "="*60)
print("DATI ESTRATTI:")
print(json.dumps(linea_data, indent=2, ensure_ascii=False)[:500] + "...")
print("="*60)

# Salva in file temporaneo per verifica
with open('linea_udine_san_daniele_temp.json', 'w', encoding='utf-8') as f:
    json.dump(linea_data, f, indent=2, ensure_ascii=False)

print("\nFile temporaneo salvato: linea_udine_san_daniele_temp.json")

