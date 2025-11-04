#!/usr/bin/env python3
# Script per leggere file Excel e vedere la struttura

from openpyxl import load_workbook
import json

wb = load_workbook('Udine San Daniele.xlsx')

print("=" * 60)
print("FOGLI TROVATI:", wb.sheetnames)
print("=" * 60)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nFOGLIO: {sheet_name}")
    print(f"   Dimensioni: {ws.max_row} righe x {ws.max_column} colonne")
    print(f"\n   Prime 30 righe:")
    print("-" * 60)
    
    for i, row in enumerate(ws.iter_rows(max_row=30), 1):
        values = [str(cell.value) if cell.value is not None else '' for cell in row]
        print(f"Riga {i:2d}: {' | '.join(values)}")

print("\n" + "=" * 60)

