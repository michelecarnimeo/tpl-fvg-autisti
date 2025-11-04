#!/usr/bin/env python3
# Script per leggere file Excel e convertire in formato database.json

from openpyxl import load_workbook
import json

wb = load_workbook('Udine San Daniele.xlsx')
ws = wb.active

# Leggi la prima riga COMPLETA (tutte le celle)
first_row = ws[1]
fermate = []
for cell in first_row:
    value = cell.value
    if value is not None and str(value).strip():
        fermate.append(str(value).strip())

print(f"Numero fermate trovate: {len(fermate)}")
print(f"Prime 5 fermate: {fermate[:5]}")
print(f"Ultime 5 fermate: {fermate[-5:]}")

# Leggi la matrice dei codici (righe 2+)
codici = []
for i in range(2, ws.max_row + 1):
    row = ws[i]
    codici_riga = []
    # Prendi tutte le celle della riga (la prima è il nome della fermata)
    for cell in row:
        value = cell.value
        if value is None or str(value).strip() == '':
            codici_riga.append('')
        else:
            codici_riga.append(str(value).strip())
    codici.append(codici_riga)

print(f"\nMatrice codici: {len(codici)} righe")
print(f"Prima riga codici (primi 10): {codici[0][:10]}")

# Funzione per generare prezzi basati sui codici tariffa
def genera_prezzi_da_codici(fermate, codici):
    """Genera prezzi basati sui codici tariffa"""
    # Mappatura codici -> prezzi base (in euro)
    codici_prezzi = {
        'U1': 1.50,
        'E1': 2.50,
        'E2': 3.50,
        'E3': 4.00,
        'E4': 4.50,
        'E5': 5.50,
        'E6': 6.50,
        'E7': 7.50
    }
    
    num_fermate = len(fermate)
    prezzi = []
    
    for i in range(num_fermate):
        riga_prezzi = []
        for j in range(num_fermate):
            if i == j:
                riga_prezzi.append(0)
            else:
                # Prendi il codice dalla matrice (i+1 perché la prima riga è header)
                if i < len(codici) and j < len(codici[i]):
                    codice = codici[i][j] if j < len(codici[i]) else ''
                else:
                    codice = ''
                
                if codice and codice in codici_prezzi:
                    prezzo = codici_prezzi[codice]
                else:
                    # Default se codice non trovato
                    prezzo = 0
                riga_prezzi.append(prezzo)
        prezzi.append(riga_prezzi)
    
    return prezzi

# Genera i prezzi
prezzi = genera_prezzi_da_codici(fermate, codici)

print(f"\nPrezzi generati: {len(prezzi)} righe x {len(prezzi[0]) if prezzi else 0} colonne")

# Genera l'oggetto linea
linea_data = {
    "nome": "Linea 401 Udine-San Daniele",
    "fermate": fermate,
    "codici": codici,
    "prezzi": prezzi
}

print("\n" + "="*60)
print("RIEPILOGO:")
print(f"Nome: {linea_data['nome']}")
print(f"Fermate: {len(linea_data['fermate'])}")
print(f"Codici: matrice {len(codici)}x{len(codici[0]) if codici else 0}")
print(f"Prezzi: matrice {len(prezzi)}x{len(prezzi[0]) if prezzi else 0}")
print("="*60)

# Salva in file temporaneo per verifica
with open('linea_udine_san_daniele_fixed.json', 'w', encoding='utf-8') as f:
    json.dump(linea_data, f, indent=2, ensure_ascii=False)

print("\nFile salvato: linea_udine_san_daniele_fixed.json")

